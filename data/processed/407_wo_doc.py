import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    """
    Converts an Excel file (.xls or .xlsx) to a CSV file by reading the contents of the Excel file
    and writing them to a new CSV file with the same name but a different extension. Allows specifying
    separate paths for the Excel file source and the CSV file destination.

    Parameters:
        file_name (str): The name of the Excel file to be converted.
        excel_file_path (str): The directory path where the Excel file is located.
        csv_file_path (str): The directory path where the CSV file should be saved.

    Returns:
        str: The name of the created CSV file.

    Requirements:
    - openpyxl.load_workbook
    - os
    - csv

    Example:
    >>> task_func('test.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    'test.csv'
    >>> task_func('nonexistent.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    Traceback (most recent call last):
       ...
    FileNotFoundError: [Errno 2] No such file or directory: '/path/to/excel/files/nonexistent.xlsx'

    Note:
    - This function assumes the active sheet is the one to be converted.
    """

    excel_file = os.path.join(excel_file_path, file_name)
    if not os.path.isfile(excel_file):
        raise FileNotFoundError(f"[Errno 2] No such file or directory: '{excel_file}'")
    workbook = load_workbook(filename=excel_file, read_only=True)
    sheet = workbook.active
    data = [[cell.value for cell in row] for row in sheet.iter_rows()]
    csv_file_name = os.path.splitext(file_name)[0] + '.csv'
    csv_file = os.path.join(csv_file_path, csv_file_name)
    with open(csv_file, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    return csv_file_name

import unittest
from unittest.mock import patch
import tempfile
import shutil
from pathlib import Path
import openpyxl
class TestCases(unittest.TestCase):
    def setUp(self):
        # Create a temporary directory
        self.test_dir = tempfile.mkdtemp()
        self.mock_excel_path = Path(self.test_dir)
        self.mock_csv_path = Path(self.test_dir)
    def tearDown(self):
        # Remove the directory after the test
        shutil.rmtree(self.test_dir)
    def create_temp_excel_file(self, file_name: str):
        """Helper function to create a temporary Excel file for testing."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Hello'
        worksheet['B1'] = 'World'
        temp_file_path = self.mock_excel_path / file_name
        workbook.save(filename=temp_file_path)
        return temp_file_path
    def test_successful_conversion(self):
        """Test that an Excel file is successfully converted to a CSV file."""
        excel_file_name = 'test.xlsx'
        self.create_temp_excel_file(excel_file_name)
        result = task_func(excel_file_name, str(self.mock_excel_path), str(self.mock_csv_path))
        self.assertEqual(result, 'test.csv')
    @patch('openpyxl.load_workbook')
    def test_return_type(self, mock_load_workbook):
        """Ensure the function returns a string indicating the CSV file name."""
        excel_file_name = 'test.xlsx'
        temp_file_path = self.create_temp_excel_file(excel_file_name)
        mock_load_workbook.return_value.active.iter_rows.return_value = iter([])
        result = task_func(excel_file_name, str(self.mock_excel_path), str(self.mock_csv_path))
        self.assertIsInstance(result, str)
    def test_file_not_found(self):
        """Check that FileNotFoundError is raised when the Excel file does not exist."""
        with self.assertRaises(FileNotFoundError):
            task_func('nonexistent.xlsx', str(self.mock_excel_path), str(self.mock_csv_path))
    def test_csv_file_creation(self):
        """Test that a CSV file is created with the expected content from the Excel file."""
        excel_file_name = 'test.xlsx'
        self.create_temp_excel_file(excel_file_name)
        # Call the function under test
        csv_file_name = task_func(excel_file_name, str(self.mock_excel_path), str(self.mock_csv_path))
        csv_file_path = self.mock_csv_path / csv_file_name
        # Check if the CSV file was actually created
        self.assertTrue(os.path.exists(csv_file_path), f"CSV file was not created: {csv_file_path}")
        # Check the content of the created CSV file
        expected_content = [['Hello', 'World']]  # Adjust this based on the actual content of your Excel file
        with open(csv_file_path, newline='', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            actual_content = list(reader)
            self.assertEqual(actual_content, expected_content, "CSV file content does not match expected content.")
